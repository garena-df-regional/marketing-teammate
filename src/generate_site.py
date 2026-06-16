"""
DF REG MKT Teammate — Site Generator
Usage: python generate_site.py
Reads the Excel file and outputs index.html in the same folder.
Re-run this script whenever the Excel content is updated.
"""
from openpyxl import load_workbook
import json, os, re, base64

EXCEL = r"C:\Users\lintim\Desktop\claude\DF REG MKT\DF REG MKT Teammate\DF Regional Marketing TeamMate.xlsx"
HERE  = os.path.dirname(os.path.abspath(__file__))

# Sidebar logo, embedded as a data URI so the page stays self-contained.
with open(os.path.join(HERE, 'assets', 'logo.png'), 'rb') as _lf:
    LOGO_DATA_URI = 'data:image/png;base64,' + base64.b64encode(_lf.read()).decode('ascii')

def c(v):
    return str(v).strip() if v is not None else ""

# ─── Sheet classification (drives the dynamic left-nav) ───────────────────────
STD = ["scenario","pic","steps","prepare","timeline","qa","links"]
SPECIAL = {'Index':'index', 'Reg & Local Contactor':'contactor', 'Social Media Link':'sm'}
SPECIAL_LABEL = {'index':'Index', 'contactor':'Contactor', 'sm':'Social Media Links'}

def slugify(name):
    s = re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_')
    return s or 'tab'

def classify(name):
    """Return (type, nav-label, nav-id) for a sheet name."""
    if name in SPECIAL:
        t = SPECIAL[name]
        return t, SPECIAL_LABEL[t], '_' + t
    m = re.match(r'^\s*Tab\s*(\d+)\s*(.*)$', name, re.I)
    if m:
        rest = m.group(2).strip()
        return 'card', ('Tab ' + m.group(1) + (' · ' + rest if rest else '')), 'tab' + m.group(1)
    return 'generic', name, 'g_' + slugify(name)

def parse_card(ws):
    rows = []
    for r in ws.iter_rows(min_row=3):
        vals = [cell.value for cell in r[:7]]
        if any(v is not None for v in vals):
            row = dict(zip(STD, [c(v) for v in vals]))
            for i, cell in enumerate(r[:7]):
                if cell.hyperlink and cell.hyperlink.target:
                    row[STD[i] + '_url'] = cell.hyperlink.target
            rows.append(row)
    return rows

def parse_index(ws):
    idx = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if any(c(v) for v in r):
            idx.append({"tab": c(r[0]), "name": c(r[1]), "desc": c(r[2])})
    return idx

def parse_contactor(ws):
    team, contactors, sec = [], [], None
    for r in ws.iter_rows(values_only=True):
        if not any(v for v in r): continue
        f = c(r[0])
        if 'Regional Team' in f:  sec = None; continue
        if f == 'Member':         sec = 'team'; continue
        if 'Local MKT' in f:     sec = None; continue
        if f == 'Region':         sec = 'loc'; continue
        if sec == 'team':
            team.append({"name": c(r[0]), "resp": c(r[1]), "email": c(r[2])})
        elif sec == 'loc':
            contactors.append({"region": c(r[0]), "contact": c(r[1]), "email": c(r[2])})
    return {"team": team, "contactors": contactors}

def parse_sm(ws):
    sm = {"off_hdr":[], "off":[], "non_hdr":[], "non":[], "esp_hdr":[], "esp":[]}
    sec = None
    for r in ws.iter_rows(values_only=True):
        first = c(r[0])
        if first == 'Official Channels': sec='off';  continue
        if first == 'Non-Official':      sec='non';  continue
        if first == 'Esports Channel':   sec='esp';  continue
        if not any(v for v in r):        continue
        if sec == 'off':
            if first == 'Channel':
                sm['off_hdr'] = [c(v) for v in r if c(v)]
            else:
                sm['off'].append([c(v) if c(v) else '—' for v in r[:len(sm['off_hdr'])]])
        elif sec == 'non':
            if first == 'Channel':
                sm['non_hdr'] = [c(v) for v in r if c(v)]
            else:
                sm['non'].append([c(v) if c(v) else '—' for v in r[:len(sm['non_hdr'])]])
        elif sec == 'esp':
            if first == 'Channel':
                sm['esp_hdr'] = [c(v) for v in r if c(v)]
            else:
                row_vals = [c(v) if c(v) else '—' for v in r]
                while row_vals and row_vals[-1] == '—': row_vals.pop()
                if any(v != '—' for v in row_vals):
                    sm['esp'].append(row_vals)
    return sm

def parse_generic(ws):
    """Any other sheet: first non-empty row = header, rest = rows (value + hyperlink per cell)."""
    headers, rows = [], []
    for r in ws.iter_rows():
        cells = list(r)
        vals = [c(cell.value) for cell in cells]
        while vals and vals[-1] == '': vals.pop()
        if not any(vals): continue
        if not headers:
            headers = vals
            continue
        rowcells = []
        for i in range(len(headers)):
            cell = cells[i] if i < len(cells) else None
            text = c(cell.value) if cell is not None else ''
            url = cell.hyperlink.target if (cell is not None and cell.hyperlink and cell.hyperlink.target) else ''
            rowcells.append({"text": text, "url": url})
        rows.append(rowcells)
    return {"headers": headers, "rows": rows}

PARSERS = {'card':parse_card, 'index':parse_index, 'contactor':parse_contactor,
           'sm':parse_sm, 'generic':parse_generic}

# ─── Build NAV + tabs dynamically, in sheet order ─────────────────────────────
wb = load_workbook(EXCEL, data_only=True)
NAV, tabs, first_card = [], {}, False
for name in wb.sheetnames:
    typ, label, tid = classify(name)
    tabs[tid] = PARSERS[typ](wb[name])
    if typ == 'card' and not first_card:   # separator before the first Tab* page
        NAV.append({"sep": True})
        first_card = True
    NAV.append({"id": tid, "label": label, "type": typ})

DATA_JSON = json.dumps({"nav": NAV, "tabs": tabs}, ensure_ascii=False)

# ─── HTML template ────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="robots" content="noindex, nofollow">
<title>DF REG MKT Teammate</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;font-size:14px;background:#f1f5f9;color:#1e293b;display:flex;height:100vh;overflow:hidden}

/* ── Sidebar ── */
#sidebar{width:228px;min-width:228px;background:#0f172a;color:#94a3b8;display:flex;flex-direction:column;overflow-y:auto;flex-shrink:0}
#sidebar-header{padding:20px 16px 14px;border-bottom:1px solid #1e293b}
#sidebar-header .logo-img{width:132px;height:auto;display:block;margin-bottom:10px}
#sidebar-header .title{font-size:15px;font-weight:700;color:#f1f5f9;line-height:1.3}
#nav-list{list-style:none;padding:8px 0;flex:1}
#nav-list li.sep{height:1px;background:#1e293b;margin:8px 12px}
#nav-list li.nav-item a{display:block;padding:8px 16px;color:#94a3b8;text-decoration:none;font-size:13px;border-radius:0;transition:background .15s,color .15s;line-height:1.4}
#nav-list li.nav-item a:hover{background:#1e293b;color:#e2e8f0}
#nav-list li.nav-item a.active{background:#1d4ed8;color:#fff;font-weight:600}

/* ── Main ── */
#main{flex:1;display:flex;flex-direction:column;overflow:hidden}
#toolbar{padding:14px 24px;background:#fff;border-bottom:1px solid #e2e8f0;display:flex;align-items:center;gap:12px;flex-shrink:0}
#toolbar h2{font-size:16px;font-weight:700;color:#0f172a;min-width:0;flex:1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
#search{padding:7px 12px;border:1px solid #cbd5e1;border-radius:6px;font-size:13px;width:260px;outline:none;background:#f8fafc}
#search:focus{border-color:#3b82f6;background:#fff}
#content{flex:1;overflow-y:auto;padding:20px 24px}

/* ── Cards ── */
.cards-grid{display:flex;flex-direction:column;gap:12px}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden;transition:box-shadow .15s}
.card:hover{box-shadow:0 2px 12px rgba(0,0,0,.07)}
.card-header{padding:14px 16px;cursor:pointer;display:flex;align-items:flex-start;gap:10px;user-select:none}
.card-header:hover{background:#f8fafc}
.card-scenario{flex:1;font-weight:600;font-size:14px;color:#0f172a;line-height:1.4;white-space:pre-line}
.card-pic{flex-shrink:0;padding:2px 9px;border-radius:20px;font-size:11px;font-weight:700;letter-spacing:.03em;margin-top:2px}
.card-chevron{color:#94a3b8;font-size:12px;margin-top:3px;flex-shrink:0;transition:transform .2s}
.card-chevron.open{transform:rotate(180deg)}
.card-body{display:none;padding:0 16px 16px;border-top:1px solid #f1f5f9}
.card-body.open{display:block}
.field{margin-top:14px}
.field-label{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#64748b;margin-bottom:5px}
.field-value{color:#1e293b;line-height:1.7;white-space:pre-wrap;word-break:break-word;font-size:13.5px}
.field-value a{color:#2563eb;text-decoration:none}
.field-value a:hover{text-decoration:underline}
.field-link{display:inline-block;margin-top:6px;padding:4px 10px;background:#eff6ff;border:1px solid #bfdbfe;border-radius:5px;color:#1d4ed8;font-size:12px;font-weight:600;text-decoration:none}
.field-link:hover{background:#dbeafe}

/* PIC colors */
.pic-tim{background:#ede9fe;color:#5b21b6}
.pic-ruru{background:#fce7f3;color:#9d174d}
.pic-jeremy{background:#fef3c7;color:#92400e}
.pic-hua{background:#d1fae5;color:#065f46}
.pic-mye{background:#dbeafe;color:#1e40af}
.pic-multi{background:#f1f5f9;color:#475569}

/* ── Tables (Index / Contactor / SM) ── */
.page-title{font-size:18px;font-weight:700;color:#0f172a;margin-bottom:18px}
.section-title{font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#64748b;margin:24px 0 10px}
.data-table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #e2e8f0;border-radius:10px;overflow:hidden}
.data-table th{background:#f8fafc;padding:10px 14px;text-align:left;font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:#64748b;border-bottom:1px solid #e2e8f0;white-space:nowrap}
.data-table td{padding:10px 14px;border-bottom:1px solid #f1f5f9;vertical-align:top;font-size:13px;line-height:1.6;word-break:break-word}
.data-table tr:last-child td{border-bottom:none}
.data-table td a{color:#2563eb;text-decoration:none}
.data-table td a:hover{text-decoration:underline}
.data-table .resp-col{white-space:pre-wrap;max-width:360px}
.sm-table-wrap{overflow-x:auto;border-radius:10px;border:1px solid #e2e8f0;margin-bottom:20px}
.sm-table{border-collapse:collapse;background:#fff;min-width:600px}
.sm-table th{background:#f8fafc;padding:9px 12px;text-align:left;font-size:12px;font-weight:700;text-transform:uppercase;color:#64748b;border-bottom:1px solid #e2e8f0;white-space:nowrap}
.sm-table td{padding:8px 12px;border-bottom:1px solid #f1f5f9;font-size:12px;word-break:break-all;max-width:180px;vertical-align:top}
.sm-table tr:last-child td{border-bottom:none}
.sm-table td:first-child{font-weight:600;white-space:nowrap;word-break:normal;max-width:none}
.sm-table td a{color:#2563eb;font-size:11px}
.dash{color:#cbd5e1}
.empty-state{padding:60px 24px;text-align:center;color:#94a3b8}
.empty-state .icon{font-size:36px;margin-bottom:12px}
.empty-state p{font-size:14px}
.no-results{padding:40px;text-align:center;color:#94a3b8;font-size:14px}
.email-link{color:#2563eb;text-decoration:none;font-size:12px}
.tag{display:inline-block;background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;border-radius:4px;padding:2px 8px;font-size:11px;font-weight:700;margin-right:4px;white-space:nowrap;min-width:30px;text-align:center}
</style>
</head>
<body>
<aside id="sidebar">
  <div id="sidebar-header">
    <img class="logo-img" src="__LOGO__" alt="Delta Force">
    <div class="title">REG MKT Teammate</div>
  </div>
  <ul id="nav-list"></ul>
</aside>
<div id="main">
  <div id="toolbar">
    <h2 id="page-title">Index</h2>
    <input id="search" type="search" placeholder="Search all content…" autocomplete="off">
  </div>
  <div id="content"></div>
</div>

<script>
const RAW = """ + DATA_JSON + r""";

// ── PIC badge helper ──────────────────────────────────────────
function picClass(pic){
  const p = (pic||'').toLowerCase();
  if(p.includes(',') || (p.includes('tim') && p.length > 5)) return 'pic-multi';
  if(p.includes('tim'))    return 'pic-tim';
  if(p.includes('ruru'))   return 'pic-ruru';
  if(p.includes('jeremy')) return 'pic-jeremy';
  if(p.includes('hua'))    return 'pic-hua';
  if(p.includes('mye'))    return 'pic-mye';
  return 'pic-multi';
}

// ── Link detector ─────────────────────────────────────────────
function linkify(text){
  if(!text) return '';
  const esc = text.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  return esc.replace(/(https?:\/\/[^\s]+)/g, '<a href="$1" target="_blank" rel="noopener">$1</a>');
}

// ── Render helpers ────────────────────────────────────────────
function field(label, value){
  if(!value || value === '—') return '';
  return `<div class="field">
    <div class="field-label">${label}</div>
    <div class="field-value">${linkify(value)}</div>
  </div>`;
}

function renderCards(rows, query){
  if(!rows || rows.length === 0)
    return `<div class="empty-state"><div class="icon">📭</div><p>No content yet — check back later.</p></div>`;
  const q = (query||'').toLowerCase();
  const filtered = rows.filter(r => {
    if(!q) return true;
    return Object.values(r).some(v => (v||'').toLowerCase().includes(q));
  });
  if(filtered.length === 0)
    return `<div class="no-results">No results for "<strong>${q}</strong>"</div>`;
  return `<div class="cards-grid">${filtered.map(r => {
    const hasLinks = r.links && r.links !== '—';
    return `<div class="card">
      <div class="card-header" onclick="toggleCard(this)">
        <div class="card-scenario">${r.scenario||'(Untitled)'}</div>
        <span class="card-pic ${picClass(r.pic)}">${r.pic||''}</span>
        <span class="card-chevron">▼</span>
      </div>
      <div class="card-body">
        ${field('Process Steps', r.steps)}
        ${field('What to Prepare', r.prepare)}
        ${field('Timeline', r.timeline)}
        ${field('Common Q&A', r.qa)}
        ${hasLinks ? `<div class="field"><div class="field-label">Related Links</div><div class="field-value">${r.links_url ? `<a href="${r.links_url}" target="_blank" rel="noopener" class="field-link">${r.links}</a>` : linkify(r.links)}</div></div>` : ''}
      </div>
    </div>`;
  }).join('')}</div>`;
}

function renderIndex(rows){
  function idForRow(r){
    const t = String(r.tab||'').trim();
    if(/^\d+$/.test(t) && RAW.nav.find(n=>n.id==='tab'+t)) return 'tab'+t;
    if(/contactor/i.test(r.name)) return '_contactor';
    if(/social media|sm link/i.test(r.name)) return '_sm';
    return null;
  }
  return `<p class="page-title">Tab Index</p>
  <table class="data-table">
    <thead><tr><th style="width:48px">Tab</th><th style="width:180px">Name</th><th>Description</th></tr></thead>
    <tbody>${rows.map(r=>{
      const tid=idForRow(r);
      const nameEl=tid
        ? `<a href="#" onclick="navigate(\'${tid}\');return false;" style="color:#2563eb;font-weight:600;text-decoration:none">${r.name} ↗</a>`
        : `<span style="font-weight:600">${r.name}</span>`;
      return `<tr>
        <td><span class="tag">${r.tab}</span></td>
        <td>${nameEl}</td>
        <td style="color:#475569">${r.desc}</td>
      </tr>`;
    }).join('')}</tbody>
  </table>`;
}

function renderContactor(data){
  const teamRows = data.team.map(r=>`<tr>
    <td style="font-weight:700">${r.name}</td>
    <td class="resp-col">${r.resp}</td>
    <td><a href="mailto:${r.email}" class="email-link">${r.email}</a></td>
  </tr>`).join('');
  const locRows = data.contactors.map(r=>`<tr>
    <td style="font-weight:700">${r.region}</td>
    <td>${r.contact}</td>
    <td><a href="mailto:${r.email}" class="email-link">${r.email}</a></td>
  </tr>`).join('');
  return `<p class="page-title">Contactor</p>
  <div class="section-title">Regional Team</div>
  <table class="data-table">
    <thead><tr><th>Member</th><th>Responsibilities</th><th>Email</th></tr></thead>
    <tbody>${teamRows}</tbody>
  </table>
  <div class="section-title">Local MKT Contactors</div>
  <table class="data-table">
    <thead><tr><th>Region</th><th>MKT Contactor</th><th>Email</th></tr></thead>
    <tbody>${locRows}</tbody>
  </table>`;
}

function smCell(val){
  if(!val || val==='—') return `<span class="dash">—</span>`;
  if(val.startsWith('http')) return `<a href="${val}" target="_blank" rel="noopener">↗ Link</a>`;
  return val;
}

function renderSM(sm){
  const offHead = sm.off_hdr.map(h=>`<th>${h}</th>`).join('');
  const offBody = sm.off.map(r=>`<tr>${r.map(v=>`<td>${smCell(v)}</td>`).join('')}</tr>`).join('');
  const nonHead = sm.non_hdr.map(h=>`<th>${h}</th>`).join('');
  const nonBody = sm.non.map(r=>`<tr>${r.map(v=>`<td>${smCell(v)}</td>`).join('')}</tr>`).join('');
  const espHead = sm.esp_hdr && sm.esp_hdr.length ? `<thead><tr>${sm.esp_hdr.map(h=>`<th>${h}</th>`).join('')}</tr></thead>` : '';
  const espSection = sm.esp.length ? `
    <div class="section-title">Esports Channels</div>
    <div class="sm-table-wrap">
      <table class="sm-table">${espHead}<tbody>${sm.esp.map(r=>`<tr>${r.map(v=>`<td>${smCell(v)}</td>`).join('')}</tr>`).join('')}</tbody>
      </table>
    </div>` : '';
  return `<p class="page-title">Social Media Links</p>
  <div class="section-title">Official Channels</div>
  <div class="sm-table-wrap">
    <table class="sm-table"><thead><tr>${offHead}</tr></thead><tbody>${offBody}</tbody></table>
  </div>
  <div class="section-title">Non-Official Channels</div>
  <div class="sm-table-wrap">
    <table class="sm-table"><thead><tr>${nonHead}</tr></thead><tbody>${nonBody}</tbody></table>
  </div>
  ${espSection}`;
}

// ── Generic table (any other sheet: 1st row = header, rest = rows) ─────────────
function genCell(cell){
  const text=(cell&&cell.text)||'', url=(cell&&cell.url)||'';
  if(url)  return `<a href="${url}" target="_blank" rel="noopener">${text||'↗ Link'}</a>`;
  if(!text) return `<span class="dash">—</span>`;
  if(/^https?:\/\//.test(text)) return `<a href="${text}" target="_blank" rel="noopener">${text}</a>`;
  return text;
}
function renderGeneric(title, data, query){
  if(!data||!data.headers||!data.headers.length)
    return `<p class="page-title">${title}</p><div class="empty-state"><div class="icon">📭</div><p>No content yet — check back later.</p></div>`;
  const q=(query||'').toLowerCase();
  const rows=data.rows.filter(row=>{
    if(!q) return true;
    return row.some(cc=>(((cc&&cc.text)||'')+' '+((cc&&cc.url)||'')).toLowerCase().includes(q));
  });
  if(!rows.length) return `<p class="page-title">${title}</p><div class="no-results">No results for "<strong>${q}</strong>"</div>`;
  const head=data.headers.map(h=>`<th>${h}</th>`).join('');
  const body=rows.map(row=>`<tr>${data.headers.map((_,i)=>`<td>${genCell(row[i])}</td>`).join('')}</tr>`).join('');
  return `<p class="page-title">${title}</p>
  <table class="data-table"><thead><tr>${head}</tr></thead><tbody>${body}</tbody></table>`;
}

// ── Navigation ────────────────────────────────────────────────
let activeId = '_index';

function getLabel(id){
  const item = RAW.nav.find(n=>n.id===id);
  return item ? item.label : id;
}

function buildNav(){
  const ul = document.getElementById('nav-list');
  RAW.nav.forEach(item => {
    const li = document.createElement('li');
    if(item.sep){ li.className='sep'; ul.appendChild(li); return; }
    li.className = 'nav-item';
    li.innerHTML = `<a href="#" data-id="${item.id}">${item.label}</a>`;
    li.querySelector('a').addEventListener('click', e => {
      e.preventDefault();
      navigate(item.id);
    });
    ul.appendChild(li);
  });
}

function navigate(id){
  activeId = id;
  document.querySelectorAll('#nav-list a').forEach(a => {
    a.classList.toggle('active', a.dataset.id === id);
  });
  document.getElementById('page-title').textContent = getLabel(id);
  document.getElementById('search').value = '';
  renderContent('');
}

function renderContent(query){
  const el = document.getElementById('content');
  const item = RAW.nav.find(n => n.id === activeId);
  const type = item ? item.type : 'card';
  if(type === 'index')     { el.innerHTML = renderIndex(RAW.tabs[activeId]); return; }
  if(type === 'contactor') { el.innerHTML = renderContactor(RAW.tabs[activeId]); return; }
  if(type === 'sm')        { el.innerHTML = renderSM(RAW.tabs[activeId]); return; }
  if(type === 'generic')   { el.innerHTML = renderGeneric(item.label, RAW.tabs[activeId], query); return; }
  el.innerHTML = renderCards(RAW.tabs[activeId] || [], query);
}

// ── Search ────────────────────────────────────────────────────
let searchTimer;
document.getElementById('search').addEventListener('input', e => {
  clearTimeout(searchTimer);
  searchTimer = setTimeout(() => renderContent(e.target.value), 120);
});

// ── Card toggle ───────────────────────────────────────────────
function toggleCard(header){
  const body    = header.nextElementSibling;
  const chevron = header.querySelector('.card-chevron');
  body.classList.toggle('open');
  chevron.classList.toggle('open');
}

// ── Init ──────────────────────────────────────────────────────
buildNav();
navigate('_index');
</script>
</body>
</html>"""

HTML = HTML.replace('__LOGO__', LOGO_DATA_URI)

out_html = os.path.join(HERE, 'index.html')
with open(out_html, 'w', encoding='utf-8') as f:
    f.write(HTML)

print(f"Generated: {out_html}")

# ─── AI-friendly plain Markdown (content.md) ──────────────────────────────────
# Same data, flattened to text so Claude/ChatGPT/Gemini can read it directly.
def build_markdown(nav, tabs):
    L = ["# DF REG MKT Teammate — Knowledge Base", ""]
    L.append("> Knowledge base for Local Marketing teams collaborating with DF Regional Marketing.")
    L.append("> Each section is a scenario: who the Regional PIC is, the process, what to prepare, timeline, Q&A, and related links.")
    L.append("")
    for item in nav:
        if item.get('sep'):
            continue
        nid, label, typ = item['id'], item['label'], item.get('type')
        if typ == 'index':
            L.append(f"## {label}")
            for r in tabs[nid]:
                L.append(f"- **{r['tab']}. {r['name']}** — {r['desc']}")
            L.append("")
        elif typ == 'contactor':
            cc = tabs[nid]
            L.append(f"## {label}")
            L.append("### Regional Team")
            for m in cc['team']:
                L.append(f"- **{m['name']}** ({m['email']}): {m['resp']}")
            L.append("### Local MKT Contactors")
            for m in cc['contactors']:
                L.append(f"- **{m['region']}**: {m['contact']} ({m['email']})")
            L.append("")
        elif typ == 'sm':
            sm = tabs[nid]
            L.append(f"## {label}")
            def sm_block(title, hdr, rows):
                if not rows:
                    return
                L.append(f"### {title}")
                for row in rows:
                    pairs = [f"{hdr[i]}: {row[i]}" for i in range(len(hdr))
                             if i < len(row) and row[i] and row[i] != '—']
                    if pairs:
                        L.append("- " + " | ".join(pairs))
            sm_block("Official Channels", sm['off_hdr'], sm['off'])
            sm_block("Non-Official Channels", sm['non_hdr'], sm['non'])
            if sm['esp']:
                L.append("### Esports Channels")
                for row in sm['esp']:
                    cells = [x for x in row if x and x != '—']
                    if cells:
                        L.append("- " + " | ".join(cells))
            L.append("")
        elif typ == 'generic':
            data = tabs[nid]
            L.append(f"## {label}")
            headers = data.get('headers', [])
            for row in data.get('rows', []):
                pairs = []
                for i, h in enumerate(headers):
                    cell = row[i] if i < len(row) else None
                    val = (cell.get('url') or cell.get('text')) if cell else ''
                    if val:
                        pairs.append(f"{h}: {val}")
                if pairs:
                    L.append("- " + " | ".join(pairs))
            L.append("")
        else:
            rows = tabs.get(nid, [])
            L.append(f"## {label}")
            if not rows:
                L.append("_(No content yet.)_")
                L.append("")
                continue
            for r in rows:
                L.append(f"### {r.get('scenario') or '(Untitled)'}")
                if r.get('pic'):      L.append(f"- **Regional PIC**: {r['pic']}")
                if r.get('steps'):    L.append(f"- **Process Steps**: {r['steps']}")
                if r.get('prepare'):  L.append(f"- **What to Prepare**: {r['prepare']}")
                if r.get('timeline'): L.append(f"- **Timeline**: {r['timeline']}")
                if r.get('qa'):       L.append(f"- **Common Q&A**: {r['qa']}")
                if r.get('links'):
                    if r.get('links_url'):
                        L.append(f"- **Related Links**: [{r['links']}]({r['links_url']})")
                    else:
                        L.append(f"- **Related Links**: {r['links']}")
                L.append("")
    return "\n".join(L)

out_md = os.path.join(HERE, 'content.md')
with open(out_md, 'w', encoding='utf-8') as f:
    f.write(build_markdown(NAV, tabs))

print(f"Generated: {out_md}")
